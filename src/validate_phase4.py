"""Phase 4 portfolio-presentation validation gate."""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from common import DATA_PROCESSED, PROJECT_ROOT, get_logger

log = get_logger("validate_phase4")
OUTPUT = DATA_PROCESSED / "analytics" / "rpt_phase4_validation.csv"
README = PROJECT_ROOT / "README.md"

DASHBOARDS = [
    "01_executive_overview.png",
    "02_shipment_control_tower.png",
    "03_carrier_lane_performance.png",
    "04_freight_audit.png",
    "05_finance_accrual.png",
    "06_data_quality_controls.png",
]
CHARTS = [
    "monthly_otif_trend.png",
    "carrier_scorecard.png",
    "lane_risk_matrix.png",
    "freight_audit_exposure_waterfall.png",
    "three_way_match_distribution.png",
    "accrual_aging.png",
    "data_quality_detection_performance.png",
    "exception_severity_distribution.png",
]
DIAGRAMS = ["project_architecture.png", "freight_audit_workflow.png", "data_quality_validation.png"]
DOCS = [
    "project_case_study.md",
    "interview_demo_script.md",
    "recruiter_walkthrough.md",
    "presentation_outline.md",
    "interview_materials.md",
    "chart_source_map.md",
    "phase4_summary.md",
]


def local_targets(markdown: str) -> list[str]:
    targets = []
    for target in re.findall(r"!?(?:\[[^\]]*\])\(([^)]+)\)", markdown):
        target = target.strip().split("#", 1)[0]
        if target and not re.match(r"^[a-z]+://", target) and not target.startswith("mailto:"):
            targets.append(target)
    return targets


def read_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def validate() -> list[dict]:
    checks = []

    def add(name, passed, detail=""):
        checks.append({"validation_name": name, "severity": "CRITICAL", "status": "PASS" if bool(passed) else "FAIL", "detail": str(detail)})

    dashboard_paths = [PROJECT_ROOT / "dashboard" / "screenshots" / name for name in DASHBOARDS]
    chart_paths = [PROJECT_ROOT / "documentation" / "charts" / name for name in CHARTS]
    diagram_paths = [PROJECT_ROOT / "documentation" / "diagrams" / name for name in DIAGRAMS]
    add("six_dashboard_pngs", len(dashboard_paths) == 6 and all(path.stat().st_size > 10_000 for path in dashboard_paths), [path.name for path in dashboard_paths])
    add("eight_standalone_chart_pngs", len(chart_paths) == 8 and all(path.stat().st_size > 10_000 for path in chart_paths), [path.name for path in chart_paths])
    add("three_diagram_pngs", len(diagram_paths) == 3 and all(path.stat().st_size > 10_000 for path in diagram_paths), [path.name for path in diagram_paths])
    add("three_mermaid_sources", all((PROJECT_ROOT / "documentation" / "diagrams" / name.replace(".png", ".mmd")).exists() for name in DIAGRAMS))

    dashboard_rows = read_csv(PROJECT_ROOT / "dashboard" / "dashboard_metric_reconciliation.csv")
    chart_rows = read_csv(PROJECT_ROOT / "documentation" / "charts" / "chart_source_reconciliation.csv")
    add("dashboard_headline_reconciliation", len(dashboard_rows) >= 30 and all(row["match"] == "PASS" for row in dashboard_rows), len(dashboard_rows))
    add("chart_source_reconciliation", len(chart_rows) == 8 and all(row["match"] == "PASS" for row in chart_rows), len(chart_rows))

    readme = README.read_text(encoding="utf-8")
    image_targets = re.findall(r"!\[[^\]]*\]\(([^)]+)\)", readme)
    add("readme_image_paths_exist", image_targets and all((PROJECT_ROOT / target).exists() for target in image_targets), image_targets)
    targets = local_targets(readme)
    missing = [target for target in targets if not (PROJECT_ROOT / target).exists()]
    add("readme_local_links_exist", not missing, missing)
    add("recruiter_first_readme", readme.startswith("# Solar Logistics Control Tower & Freight Audit System") and "## Key outcomes" in readme and "## Executive dashboard preview" in readme)
    add("readme_financial_disclosure", "modeled within simulated enterprise records" in readme.lower())

    doc_paths = [PROJECT_ROOT / "documentation" / name for name in DOCS]
    add("required_phase4_documentation", all(path.exists() and path.stat().st_size > 500 for path in doc_paths), [path.name for path in doc_paths])
    disclosure_docs = [README, PROJECT_ROOT / "documentation" / "project_case_study.md", PROJECT_ROOT / "documentation" / "recruiter_walkthrough.md", PROJECT_ROOT / "documentation" / "interview_demo_script.md"]
    add("documentation_disclosure_coverage", all("simulat" in path.read_text(encoding="utf-8").lower() and "public" in path.read_text(encoding="utf-8").lower() for path in disclosure_docs))

    wb = load_workbook(PROJECT_ROOT / "excel" / "logistics_kpi_pack.xlsx", read_only=True, data_only=False)
    expected = ["Executive Summary", "Shipment Exceptions", "Carrier Scorecard", "Lane Scorecard", "Freight Audit", "Three-Way Match", "Accrual Report", "Open Claims", "Data Quality", "Metric Definitions"]
    add("excel_expected_worksheets", wb.sheetnames == expected, wb.sheetnames)
    ws = wb["Executive Summary"]
    metrics = {ws.cell(row, 1).value: ws.cell(row, 2).value for row in range(7, ws.max_row + 1)}
    add("excel_headlines_reconcile", metrics.get("Shipment count") == 10324 and metrics.get("Recoverable overcharge") == 428053.23 and metrics.get("Critical recall") == 100)
    wb.close()
    excel_report = (PROJECT_ROOT / "excel" / "kpi_pack_validation.md").read_text(encoding="utf-8")
    add("excel_validation_report", "44/44 checks passed" in excel_report)

    hash_manifest = PROJECT_ROOT / "documentation" / "phase4_artifact_hashes.csv"
    hash_rows = read_csv(hash_manifest) if hash_manifest.exists() else []
    add("phase4_idempotency_manifest", len(hash_rows) >= 20 and all(row["match"] == "PASS" for row in hash_rows), len(hash_rows))
    return checks


def main() -> int:
    checks = validate()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["validation_name", "severity", "status", "detail"], lineterminator="\n")
        writer.writeheader()
        writer.writerows(checks)
    failures = [check for check in checks if check["status"] == "FAIL"]
    if failures:
        for check in failures:
            log.error("Phase 4 validation failed: %s | %s", check["validation_name"], check["detail"])
        return 1
    log.info("Phase 4 validation: %d/%d checks pass", len(checks), len(checks))
    return 0


if __name__ == "__main__":
    sys.exit(main())
