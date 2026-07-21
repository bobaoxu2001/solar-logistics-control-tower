"""Publication-safety checks for tracked repository content and assets."""

from __future__ import annotations

import re
import subprocess
import sys
from pathlib import Path
from urllib.parse import unquote

from openpyxl import load_workbook

from common import PROJECT_ROOT, get_logger

log = get_logger("check_repository_health")

TEXT_SUFFIXES = {
    "", ".csv", ".dax", ".env", ".gitignore", ".json", ".md", ".mmd",
    ".py", ".sql", ".toml", ".txt", ".yaml", ".yml",
}

LOCAL_ROOT_PATTERNS = [
    ("macOS user directory", re.compile(re.escape("/" + "Users" + "/"))),
    ("Linux user directory", re.compile(re.escape("/" + "home" + "/"))),
    ("Windows drive path", re.compile(r"(?<![A-Za-z0-9])[A-Za-z]:[\\/]")),
    ("local username", re.compile("xu" + "ao", re.IGNORECASE)),
    ("personal documents path", re.compile("Documents" + r"[\\/]", re.IGNORECASE)),
    ("personal workspace folder", re.compile("2025" + r"\s+" + "找工作")),
]

KEY_PATTERNS = [
    ("private key", re.compile(r"BEGIN (?:RSA |EC |OPENSSH )?PRIVATE KEY")),
    ("AWS access key", re.compile(r"AKIA[0-9A-Z]{16}")),
    ("GitHub token", re.compile(r"gh[pousr]_[A-Za-z0-9_]{20,}")),
    ("OpenAI-style key", re.compile(r"sk-[A-Za-z0-9_-]{20,}")),
    ("Slack token", re.compile(r"xox[baprs]-[A-Za-z0-9-]{20,}")),
]
ASSIGNMENT = re.compile(
    r"(?i)\b(password|passwd|token|secret|api[_-]?key|client[_-]?secret)\b\s*[:=]\s*[\"']?([^\"'\s,}\]]+)"
)
AUTH_URL = re.compile(r"[a-z][a-z0-9+.-]*://[^\s/:]+:([^\s/@]+)@", re.IGNORECASE)
EMAIL = re.compile(r"(?i)\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b")
PHONE = re.compile(r"(?<!\w)(?:\+\d{1,3}[ .-]?)?(?:\(\d{3}\)|\d{3})[ .-]\d{3}[ .-]\d{4}(?!\w)")
PLACEHOLDERS = (
    "change-me", "local-only", "placeholder", "example", "dummy", "not-set",
    "replace-me", "test-only", "sunlog" + "_dev_" + "password",
)

REQUIRED_DASHBOARDS = [
    "01_executive_overview.png", "02_shipment_control_tower.png",
    "03_carrier_lane_performance.png", "04_freight_audit.png",
    "05_finance_accrual.png", "06_data_quality_controls.png",
]
REQUIRED_DIAGRAMS = [
    "project_architecture.png", "freight_audit_workflow.png",
    "data_quality_validation.png",
]
REQUIRED_FILES = [
    ".github/workflows/tests.yml", "LICENSE", "DATA_NOTICE.md",
    "README.md", "PORTFOLIO_GUIDE.md", "dashboard/measures.dax",
    "dashboard/dashboard_metric_reconciliation.csv",
    "documentation/phase2_summary.md", "documentation/phase3_summary.md",
    "documentation/phase4_summary.md", "documentation/project_case_study.md",
    "documentation/github_release_checklist.md", "documentation/github_metadata.md",
    "documentation/publication_audit_summary.md",
    "documentation/charts/chart_source_reconciliation.csv",
    "excel/logistics_kpi_pack.xlsx", "excel/kpi_pack_validation.md",
]


def git(*args: str, check: bool = True) -> subprocess.CompletedProcess:
    return subprocess.run(
        ["git", *args], cwd=PROJECT_ROOT, check=check, capture_output=True, text=True
    )


def tracked_files(revision: str | None = None) -> list[str]:
    if revision:
        result = git("ls-tree", "-r", "--name-only", revision)
    else:
        # Include intended, non-ignored publication files before they are committed.
        result = git("ls-files", "--cached", "--others", "--exclude-standard")
    return [line for line in result.stdout.splitlines() if line]


def decode_text(data: bytes, path: str) -> str | None:
    if Path(path).suffix.lower() not in TEXT_SUFFIXES and Path(path).name not in {".env.example", ".gitignore"}:
        return None
    if b"\x00" in data:
        return None
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError:
        return None


def tracked_texts(revision: str | None = None):
    for path in tracked_files(revision):
        if revision:
            result = subprocess.run(
                ["git", "show", f"{revision}:{path}"], cwd=PROJECT_ROOT,
                check=False, capture_output=True,
            )
            if result.returncode:
                continue
            data = result.stdout
        else:
            file_path = PROJECT_ROOT / path
            if not file_path.is_file():
                continue
            data = file_path.read_bytes()
        text = decode_text(data, path)
        if text is not None:
            yield path, text


def scan_local_paths(revision: str | None = None) -> list[dict]:
    issues = []
    for path, text in tracked_texts(revision):
        for line_number, line in enumerate(text.splitlines(), 1):
            for name, pattern in LOCAL_ROOT_PATTERNS:
                if pattern.search(line):
                    issues.append({"kind": name, "path": path, "line": line_number})
    return issues


def is_placeholder(value: str) -> bool:
    normalized = value.lower()
    return any(marker in normalized for marker in PLACEHOLDERS) or normalized in {"", "none", "null"}


def scan_secrets(revision: str | None = None) -> list[dict]:
    issues = []
    for path, text in tracked_texts(revision):
        for line_number, line in enumerate(text.splitlines(), 1):
            for kind, pattern in KEY_PATTERNS:
                if pattern.search(line):
                    issues.append({"kind": kind, "path": path, "line": line_number})
            for match in ASSIGNMENT.finditer(line):
                value = match.group(2)
                if not is_placeholder(value) and not value.startswith("${"):
                    issues.append({"kind": "credential assignment", "path": path, "line": line_number})
            for match in AUTH_URL.finditer(line):
                if not is_placeholder(match.group(1)):
                    issues.append({"kind": "credential in URL", "path": path, "line": line_number})
    return issues


def scan_personal_info(revision: str | None = None) -> list[dict]:
    issues = []
    for path, text in tracked_texts(revision):
        for line_number, line in enumerate(text.splitlines(), 1):
            if EMAIL.search(line):
                issues.append({"kind": "email address", "path": path, "line": line_number})
            if PHONE.search(line):
                issues.append({"kind": "phone number", "path": path, "line": line_number})
    return issues


def markdown_files() -> list[Path]:
    return sorted(
        PROJECT_ROOT / path
        for path in tracked_files()
        if Path(path).suffix.lower() == ".md" and (PROJECT_ROOT / path).is_file()
    )


def markdown_targets(text: str) -> list[str]:
    return [match.strip() for match in re.findall(r"!?\[[^\]]*\]\(([^)]+)\)", text)]


def has_exact_case(path: Path) -> bool:
    """Check every repository-relative component without trusting a case-folding FS."""
    try:
        relative = path.resolve().relative_to(PROJECT_ROOT.resolve())
    except ValueError:
        return False
    parent = PROJECT_ROOT.resolve()
    for part in relative.parts:
        if part not in {entry.name for entry in parent.iterdir()}:
            return False
        parent /= part
    return True


def is_git_ignored(path: Path) -> bool:
    try:
        relative = path.resolve().relative_to(PROJECT_ROOT.resolve())
    except ValueError:
        return False
    result = git("check-ignore", "--quiet", "--", str(relative), check=False)
    return result.returncode == 0


def scan_markdown_links() -> list[dict]:
    issues = []
    for markdown in markdown_files():
        for target in markdown_targets(markdown.read_text(encoding="utf-8")):
            clean = unquote(target.split("#", 1)[0].strip().strip("<>"))
            if not clean or clean.startswith(("http://", "https://", "mailto:")):
                continue
            if Path(clean).is_absolute():
                issues.append({"kind": "absolute link", "path": str(markdown.relative_to(PROJECT_ROOT)), "target": clean})
                continue
            resolved = (markdown.parent / clean).resolve()
            if not resolved.exists():
                issues.append({"kind": "broken link", "path": str(markdown.relative_to(PROJECT_ROOT)), "target": clean})
            elif PROJECT_ROOT.resolve() not in [resolved, *resolved.parents]:
                issues.append({"kind": "link outside repository", "path": str(markdown.relative_to(PROJECT_ROOT)), "target": clean})
            elif not has_exact_case(resolved):
                issues.append({"kind": "case mismatch", "path": str(markdown.relative_to(PROJECT_ROOT)), "target": clean})
            elif is_git_ignored(resolved):
                issues.append({"kind": "linked file ignored", "path": str(markdown.relative_to(PROJECT_ROOT)), "target": clean})
    return issues


def required_asset_issues() -> list[dict]:
    issues = []
    paths = [PROJECT_ROOT / path for path in REQUIRED_FILES]
    paths += [PROJECT_ROOT / "dashboard" / "screenshots" / name for name in REQUIRED_DASHBOARDS]
    paths += [PROJECT_ROOT / "documentation" / "diagrams" / name for name in REQUIRED_DIAGRAMS]
    for path in paths:
        if not path.exists() or (path.is_file() and path.stat().st_size == 0):
            issues.append({"kind": "missing required asset", "path": str(path.relative_to(PROJECT_ROOT))})
        elif path.suffix.lower() == ".png":
            data = path.read_bytes()
            if not (data.startswith(b"\x89PNG\r\n\x1a\n") and b"IHDR" in data[:32] and data[-8:-4] == b"IEND"):
                issues.append({"kind": "invalid PNG asset", "path": str(path.relative_to(PROJECT_ROOT))})
    return issues


def workbook_issues() -> list[dict]:
    path = PROJECT_ROOT / "excel" / "logistics_kpi_pack.xlsx"
    if not path.exists():
        return [{"kind": "missing workbook", "path": str(path.relative_to(PROJECT_ROOT))}]
    expected = [
        "Executive Summary", "Shipment Exceptions", "Carrier Scorecard",
        "Lane Scorecard", "Freight Audit", "Three-Way Match", "Accrual Report",
        "Open Claims", "Data Quality", "Metric Definitions",
    ]
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        issues = [] if workbook.sheetnames == expected else [{"kind": "unexpected workbook sheets", "path": str(path.relative_to(PROJECT_ROOT))}]
        properties = workbook.properties
        metadata = " ".join(
            str(getattr(properties, field) or "")
            for field in ["creator", "lastModifiedBy", "title", "subject", "description", "keywords", "category"]
        )
        if EMAIL.search(metadata) or any(pattern.search(metadata) for _, pattern in LOCAL_ROOT_PATTERNS):
            issues.append({"kind": "personal workbook metadata", "path": str(path.relative_to(PROJECT_ROOT))})
        workbook.close()
        return issues
    except Exception:  # noqa: BLE001 - report integrity failure without leaking internals
        return [{"kind": "workbook integrity failure", "path": str(path.relative_to(PROJECT_ROOT))}]


def disclosure_issues() -> list[dict]:
    readme = (PROJECT_ROOT / "README.md").read_text(encoding="utf-8").lower()
    required = {
        "public shipment disclosure": "public shipment history",
        "simulated enterprise disclosure": "simulated enterprise records",
        "modeled exposure disclosure": "modeled within simulated enterprise records",
        "non-affiliation disclosure": "not affiliated",
        "source-terms caveat": "exact reuse terms were not independently verified",
        "checksum disclosure": "918b992d",
    }
    return [
        {"kind": "missing disclosure", "path": "README.md", "detail": name}
        for name, phrase in required.items() if phrase not in readme
    ]


def tracked_env_issues() -> list[dict]:
    return [
        {"kind": "tracked environment file", "path": path}
        for path in tracked_files()
        if Path(path).name.startswith(".env") and Path(path).name != ".env.example"
    ]


def run_health_checks() -> list[dict]:
    categories = [
        ("Required files and assets", required_asset_issues()),
        ("Markdown links", scan_markdown_links()),
        ("Portable tracked paths", scan_local_paths()),
        ("Credential patterns", scan_secrets()),
        ("Personal contact patterns", scan_personal_info()),
        ("Tracked environment files", tracked_env_issues()),
        ("Workbook integrity", workbook_issues()),
        ("Data and exposure disclosure", disclosure_issues()),
    ]
    return [
        {"check": name, "status": "PASS" if not issues else "FAIL", "issue_count": len(issues), "issues": issues}
        for name, issues in categories
    ]


def main() -> int:
    checks = run_health_checks()
    failures = [check for check in checks if check["status"] == "FAIL"]
    for check in checks:
        log.info("%s: %s (%d issues)", check["check"], check["status"], check["issue_count"])
    if failures:
        for check in failures:
            for issue in check["issues"]:
                log.error("%s issue in %s at line %s", issue["kind"], issue.get("path", "unknown"), issue.get("line", "n/a"))
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
