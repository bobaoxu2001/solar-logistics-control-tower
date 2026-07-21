"""Generate the Phase 3 Excel KPI pack from live database results using openpyxl."""

from __future__ import annotations

import math
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule, IconSetRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common import PROJECT_ROOT, create_project_engine, get_logger, load_config

log = get_logger("build_excel_kpi_pack")
OUTPUT = PROJECT_ROOT / "excel" / "logistics_kpi_pack.xlsx"

NAVY = "17365D"
BLUE = "2F75B5"
TEAL = "0F766E"
LIGHT_BLUE = "DCE6F1"
LIGHT_GRAY = "E7E6E6"
WHITE = "FFFFFF"
GREEN = "C6EFCE"
AMBER = "FFEB9C"
RED = "FFC7CE"
THIN_GRAY = Side(style="thin", color="D9E2F3")


def _clean(value):
    if value is None:
        return None
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "item"):
        return value.item()
    return value


def _query(engine, sql: str) -> pd.DataFrame:
    return pd.read_sql(sql, engine)


def executive_metrics(engine) -> pd.DataFrame:
    ship = _query(engine, "SELECT * FROM v_kpi_otif_summary").iloc[0]
    git = _query(engine, "SELECT * FROM v_kpi_git_summary").iloc[0]
    freight = _query(engine, "SELECT * FROM v_kpi_freight_summary").iloc[0]
    pod = _query(engine, "SELECT * FROM v_kpi_pod_compliance").iloc[0]
    claims = _query(engine, "SELECT * FROM v_kpi_claims").iloc[0]
    accrual = _query(engine, "SELECT * FROM v_accrual_summary").iloc[0]
    recover = _query(engine, """SELECT ROUND(overcharge_recoverable, 2) AS overcharge_recoverable,
                                       ROUND(duplicate_invoice_exposure, 2) AS duplicate_invoice_exposure,
                                       ROUND(accessorial_recoverable, 2) AS accessorial_recoverable
                                  FROM v_audit_recoverable_summary""").iloc[0]
    audit = _query(engine, "SELECT COUNT(*) invoice_count, SUM(CASE WHEN audit_status='MATCHED' THEN 1 ELSE 0 END) matched_count, SUM(expected_total) expected_total FROM v_freight_audit").iloc[0]
    transit = _query(engine, "SELECT AVG(actual_transit_days) avg_transit_days FROM analytics_shipment WHERE is_delivered=1").iloc[0]
    dq = _query(engine, "SELECT SUM(true_positive_count) tp, SUM(false_positive_count) fp, SUM(false_negative_count) fn FROM rpt_dq_detection_performance WHERE manifest_count>0").iloc[0]
    crit = _query(engine, "SELECT SUM(true_positive_count) tp, SUM(false_negative_count) fn FROM rpt_dq_detection_performance WHERE manifest_count>0 AND severity='CRITICAL'").iloc[0]
    weighted = _query(engine, "SELECT SUM(w.weight) weighted_open FROM dq_detected_exception d JOIN meta_dq_severity_weight w ON d.severity=w.severity WHERE d.resolution_status='OPEN'").iloc[0]
    dq_score = max(0.0, 1 - float(weighted.weighted_open) / (float(ship.shipment_count) * 10.0)) * 100
    precision = float(dq.tp) / (float(dq.tp) + float(dq.fp)) * 100
    recall = float(dq.tp) / (float(dq.tp) + float(dq.fn)) * 100
    critical_recall = float(crit.tp) / (float(crit.tp) + float(crit.fn)) * 100
    invoice_accuracy = float(audit.matched_count) / float(audit.invoice_count) * 100
    overcharge_recoverable = float(recover.overcharge_recoverable or 0)
    duplicate_exposure = float(recover.duplicate_invoice_exposure or 0)
    accessorial_recoverable = float(recover.accessorial_recoverable or 0)
    total_recoverable = overcharge_recoverable + duplicate_exposure + accessorial_recoverable

    rows = [
        ("Shipment count", ship.shipment_count, "count", "Informational", "Distinct source-derived shipments"),
        ("Delivered shipments", ship.delivered_count, "count", "Informational", "Shipments delivered by the reporting snapshot"),
        ("Goods in transit", ship.in_transit_count, "count", "Informational", "Phase 2 snapshot population; future modeled departures age at zero"),
        ("OTIF", ship.otif_pct, "percent", "Pass" if ship.otif_pct >= 85 else "Action", "Delivered on/before plan and in full"),
        ("On-time rate", ship.on_time_pct, "percent", "Pass" if ship.on_time_pct >= 85 else "Action", "Delivered on/before planned delivery date"),
        ("In-full rate", ship.in_full_pct, "percent", "Pass" if ship.in_full_pct >= 95 else "Action", "Delivered quantity at least planned quantity"),
        ("Average transit time", transit.avg_transit_days, "days", "Informational", "Mean actual ship-to-delivery days for delivered shipments"),
        ("Average delay", ship.avg_delay_days, "days", "Informational", "Actual minus planned delivery days; negative is early"),
        ("GIT value", git.git_value_usd, "USD", "Informational", "Shipment value for snapshot GIT population"),
        ("Overdue GIT", git.overdue_git_count, "count", "Action" if git.overdue_git_count else "Pass", "GIT with planned delivery before snapshot"),
        ("Freight spend", freight.invoiced_freight_usd, "USD", "Informational", "Valid-shipment invoices as billed; includes controlled exceptions"),
        ("Expected freight", audit.expected_total, "USD", "Informational", "Rated expected total; unrated invoices remain unknown"),
        ("Recoverable overcharge", overcharge_recoverable, "USD", "Action", "Material modeled overcharge and incorrect-fuel exposure"),
        ("Total recoverable exposure", total_recoverable, "USD", "Action", "Overcharge + duplicate invoice + unauthorized/excessive accessorial exposure"),
        ("Invoice accuracy", invoice_accuracy, "percent", "Action" if invoice_accuracy < 95 else "Pass", "Share of audit rows classified MATCHED"),
        ("Open accrual", accrual.open_accrual_balance, "USD", "Informational", "Expected freight on accruals not released"),
        ("Claims rate", claims.claims_rate_pct, "percent", "Watch" if claims.claims_rate_pct >= 2 else "Pass", "Claims per delivered shipment"),
        ("POD compliance", pod.pod_compliance_pct, "percent", "Pass" if pod.pod_compliance_pct >= 95 else "Action", "Delivered shipments with proof of delivery"),
        ("Data-quality score", dq_score, "percent", "Watch" if dq_score < 95 else "Pass", "1 - weighted open exceptions / maximum shipment exposure"),
        ("Detection precision", precision, "percent", "Informational", "Manifest-backed true positives divided by all detections"),
        ("Detection recall", recall, "percent", "Pass" if recall >= 95 else "Action", "Manifest-backed true positives divided by manifested records"),
        ("Critical recall", critical_recall, "percent", "Pass" if critical_recall == 100 else "Action", "Recall across critical manifested exception types"),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value", "Unit", "Status", "Definition"])


def _style_sheet(ws, title: str, span_columns: int, disclosure: bool = False):
    ws.sheet_view.showGridLines = False
    span = get_column_letter(max(1, span_columns))
    ws.merge_cells(f"A1:{span}1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A2:{span}2")
    ws["A2"] = "Snapshot: 2025-07-01 | Refresh: run python src/run_phase3.py"
    ws["A2"].font = Font(name="Aptos", size=9, italic=True, color="44546A")
    if disclosure:
        ws.merge_cells(f"A3:{span}4")
        ws["A3"] = ("Disclosure: shipment patterns originate from a public USAID dataset. Solar product mapping and "
                    "enterprise carrier, invoice, rate, milestone, claim, capacity, approval, and accrual records are "
                    "derived or simulated for portfolio analysis; results are not real corporate operations or budgets.")
        ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        ws["A3"].fill = PatternFill("solid", fgColor="FFF2CC")


def write_frame(ws, df: pd.DataFrame, start_row: int = 6):
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(name="Aptos", bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=NAVY))
    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, _clean(value))
            cell.font = Font(name="Aptos", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = Border(bottom=THIN_GRAY)
            name = headers[c_idx - 1].lower()
            if "date" in name or "timestamp" in name:
                if isinstance(cell.value, str) and cell.value:
                    parsed = pd.to_datetime(cell.value, errors="coerce")
                    if not pd.isna(parsed):
                        cell.value = parsed.to_pydatetime()
                cell.number_format = "yyyy-mm-dd"
            elif ("_usd" in name or any(k in name for k in ["amount", "cost", "spend", "freight", "charge", "accrual", "exposure"])
                  or name in {"expected_total", "invoiced_total", "variance_amount", "overcharge_amount", "undercharge_amount", "tax_amount"}):
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00'
            elif "pct" in name or "percent" in name or "rate" in name:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00"%"'
            elif isinstance(cell.value, int):
                cell.number_format = "#,##0"
            elif isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
    last_row = start_row + len(df)
    last_col = len(headers)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = f"A{start_row + 1}"
    ws.row_dimensions[start_row].height = 30
    for idx, header in enumerate(headers, 1):
        sample = [str(header)] + [str(_clean(v) or "") for v in df.iloc[:200, idx - 1].tolist()]
        width = min(max(len(v) for v in sample) + 2, 45)
        ws.column_dimensions[get_column_letter(idx)].width = max(width, 11)
    return last_row, last_col


def _text_rule(ws, col_letter: str, start: int, end: int, text: str, color: str):
    ws.conditional_formatting.add(
        f"{col_letter}{start}:{col_letter}{end}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("{text}",{col_letter}{start}))'],
                    fill=PatternFill("solid", fgColor=color)))


def metric_definitions() -> pd.DataFrame:
    rows = [
        ("OTIF", "Delivered on/before planned date AND delivered quantity >= planned quantity", "Delivered shipment", "Derived"),
        ("GIT", "Shipment status is IN_TRANSIT at the configured snapshot", "Shipment", "Derived from public pattern"),
        ("GIT age", "max(snapshot - modeled actual ship date, 0)", "GIT shipment", "Derived"),
        ("Freight spend", "Invoice total for invoices joined to a valid shipment", "Invoice", "Simulated"),
        ("Expected freight", "max(rate/kg * weight, minimum charge) + fuel + allowed supported accessorial + tax", "Rated invoice", "Simulated"),
        ("Recoverable overcharge", "Positive material expected-vs-invoiced variance for overcharge/fuel statuses", "Invoice", "Simulated control exposure"),
        ("Invoice accuracy", "Invoices with MATCHED audit status / all audit rows", "Invoice", "Derived"),
        ("Open accrual", "Expected freight where accrual status is not RELEASED", "Accrual", "Derived; not corporate budget"),
        ("Claims rate", "Claims / delivered shipments", "Claim and shipment", "Simulated"),
        ("POD compliance", "Delivered shipments with POD / delivered shipments", "Shipment", "Simulated control"),
        ("DQ score", "1 - severity-weighted open exceptions / (shipment count * critical weight)", "Detected exception", "Derived"),
        ("Precision", "True positives / (true positives + false positives), manifest-backed types", "Exception type", "Derived"),
        ("Recall", "True positives / (true positives + false negatives), manifest-backed types", "Exception type", "Derived"),
        ("Carrier score", "Weighted min-max peer score; insufficient-volume carriers are not ranked", "Carrier", "Derived"),
        ("Lane classification", "Config volume gate plus peer p75 cost/variability/accessorial and OTIF <80% rules", "Lane", "Derived"),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Definition", "Grain", "Data classification"])


def main() -> int:
    cfg = load_config()
    engine = create_project_engine(cfg)
    frames = {
        "Executive Summary": executive_metrics(engine),
        "Shipment Exceptions": _query(engine, "SELECT * FROM rpt_fact_shipment WHERE late_flag=1 OR partial_flag=1 OR overdue_git_flag=1 OR (is_delivered=1 AND has_pod=0) ORDER BY shipment_id"),
        "Carrier Scorecard": _query(engine, "SELECT * FROM rpt_carrier_scorecard ORDER BY carrier_rank, carrier_id"),
        "Lane Scorecard": _query(engine, "SELECT * FROM rpt_lane_scorecard ORDER BY insufficient_volume_flag, service_classification, lane_id"),
        "Freight Audit": _query(engine, "SELECT * FROM rpt_fact_freight_audit WHERE audit_status<>'MATCHED' ORDER BY audit_status, invoice_id"),
        "Three-Way Match": _query(engine, "SELECT * FROM v_three_way_match ORDER BY overall_match_status, invoice_id"),
        "Accrual Report": _query(engine, "SELECT * FROM rpt_fact_accrual ORDER BY accounting_period, accrual_id"),
        "Open Claims": _query(engine, "SELECT * FROM rpt_fact_claim WHERE claim_status='OPEN' ORDER BY created_date, claim_id"),
        "Data Quality": _query(engine, "SELECT * FROM rpt_fact_data_quality ORDER BY severity, business_owner, detected_exception_id"),
        "Metric Definitions": metric_definitions(),
    }
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, df in frames.items():
        ws = wb.create_sheet(sheet_name)
        _style_sheet(ws, sheet_name, min(max(len(df.columns), 4), 12),
                     disclosure=(sheet_name in {"Executive Summary", "Metric Definitions"}))
        last_row, _ = write_frame(ws, df, start_row=6)
        headers = {cell.value: cell.column_letter for cell in ws[6]}
        if sheet_name == "Executive Summary":
            status_col = headers["Status"]
            _text_rule(ws, status_col, 7, last_row, "Pass", GREEN)
            _text_rule(ws, status_col, 7, last_row, "Watch", AMBER)
            _text_rule(ws, status_col, 7, last_row, "Action", RED)
            for row in range(7, last_row + 1):
                unit = ws.cell(row, 3).value
                value_cell = ws.cell(row, 2)
                if unit == "USD":
                    value_cell.number_format = '"$"#,##0.00'
                elif unit == "percent":
                    value_cell.number_format = '0.00"%"'
                elif unit == "count":
                    value_cell.number_format = "#,##0"
                else:
                    value_cell.number_format = "#,##0.00"
        if sheet_name == "Metric Definitions":
            ws.column_dimensions["B"].width = 85
            ws.column_dimensions["C"].width = 28
            ws.column_dimensions["D"].width = 28
            for row in range(7, last_row + 1):
                ws.cell(row, 2).alignment = Alignment(vertical="top", wrap_text=True)
                ws.row_dimensions[row].height = 30
        for header, col in headers.items():
            label = str(header).lower()
            if label in {"classification", "service_classification", "audit_status", "overall_match_status", "severity"}:
                _text_rule(ws, col, 7, last_row, "Preferred", GREEN)
                _text_rule(ws, col, 7, last_row, "Stable", GREEN)
                _text_rule(ws, col, 7, last_row, "MATCHED", GREEN)
                _text_rule(ws, col, 7, last_row, "Acceptable", LIGHT_BLUE)
                _text_rule(ws, col, 7, last_row, "WARNING", AMBER)
                _text_rule(ws, col, 7, last_row, "Improvement", AMBER)
                _text_rule(ws, col, 7, last_row, "HIGH", AMBER)
                _text_rule(ws, col, 7, last_row, "CRITICAL", RED)
                _text_rule(ws, col, 7, last_row, "BLOCK", RED)
                _text_rule(ws, col, 7, last_row, "OVERCHARGE", RED)
            if label in {"total_score", "otif_pct", "invoice_accuracy_pct", "pod_compliance_pct"}:
                ws.conditional_formatting.add(
                    f"{col}7:{col}{last_row}",
                    IconSetRule(icon_style="3TrafficLights1", type="num", values=[0, 50, 80], showValue=True))
            if "variance_amount" in label or "overcharge_amount" in label:
                ws.conditional_formatting.add(
                    f"{col}7:{col}{last_row}",
                    CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=RED)))
            if any(k in label for k in ["description", "definition", "recommended_action", "exception_reason"]):
                ws.column_dimensions[col].width = 45
                for row in range(7, min(last_row, 250) + 1):
                    ws.cell(row, ws[col + "1"].column).alignment = Alignment(vertical="top", wrap_text=True)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    log.info("Wrote %s with %d sheets", OUTPUT, len(wb.sheetnames))
    return 0


if __name__ == "__main__":
    sys.exit(main())
