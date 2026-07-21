"""Phase 3 validation gate across SQL, reporting exports, and Excel."""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import inspect

from common import DATA_PROCESSED, PROJECT_ROOT, create_project_engine, get_logger, load_config
from gen_common import replace_table

log = get_logger("validate_phase3")
REPORTING = DATA_PROCESSED / "reporting"
WORKBOOK = PROJECT_ROOT / "excel" / "logistics_kpi_pack.xlsx"


def main() -> int:
    cfg = load_config()
    engine = create_project_engine(cfg)
    schema = cfg["database"]["schema"]
    insp = inspect(engine)
    tables = set(insp.get_table_names(schema=None if engine.dialect.name == "sqlite" else schema))
    views = set(insp.get_view_names(schema=None if engine.dialect.name == "sqlite" else schema))
    checks = []

    def add(name, passed, severity="CRITICAL", detail=""):
        checks.append({"validation_name": name, "severity": severity,
                       "status": "PASS" if bool(passed) else "FAIL", "detail": str(detail)})

    required_views = {
        "v_kpi_otif_summary", "v_kpi_git_summary", "v_freight_audit", "v_three_way_match",
        "v_accrual_summary", "v_carrier_metrics", "v_lane_metrics", "rpt_fact_shipment",
        "rpt_fact_freight_audit", "rpt_carrier_scorecard", "rpt_lane_scorecard",
    }
    add("required_reporting_views", required_views <= views, detail=sorted(required_views - views))
    required_tables = {"analytics_shipment", "dq_detected_exception", "carrier_scorecard_result",
                       "lane_scorecard_result", "root_cause_case_study", "reporting_date_dim"}
    add("required_phase3_tables", required_tables <= tables, detail=sorted(required_tables - tables))

    ship = pd.read_sql("SELECT * FROM analytics_shipment", engine)
    perf = pd.read_sql("SELECT * FROM rpt_dq_detection_performance WHERE manifest_count>0", engine)
    overall_recall = perf["true_positive_count"].sum() / (perf["true_positive_count"].sum() + perf["false_negative_count"].sum())
    critical = perf[perf["severity"] == "CRITICAL"]
    critical_recall = critical["true_positive_count"].sum() / (critical["true_positive_count"].sum() + critical["false_negative_count"].sum())
    add("overall_detection_recall_at_least_95", overall_recall >= 0.95, detail=f"{overall_recall:.6f}")
    add("critical_detection_recall_100", critical_recall == 1.0, detail=f"{critical_recall:.6f}")
    add("git_age_nonnegative", ship.loc[ship["git_flag"] == 1, "git_age_days"].min() >= 0)
    add("shipment_count_10324", len(ship) == 10324, detail=len(ship))
    add("full_shipment_lineage", ship["source_record_id"].nunique() == 10324 and ship["source_record_id"].notna().all())

    weight_sum = pd.read_sql("SELECT SUM(weight) total FROM meta_scorecard_weight", engine).iloc[0, 0]
    add("carrier_weights_sum_to_one", abs(weight_sum - 1.0) < 1e-9, detail=weight_sum)
    carrier = pd.read_sql("SELECT * FROM carrier_scorecard_result", engine)
    min_carrier = int(cfg["phase3"]["carrier_scorecard"]["min_shipments"])
    add("insufficient_carriers_unranked",
        carrier.loc[carrier["shipment_count"] < min_carrier, "carrier_rank"].isna().all())
    add("eligible_carrier_ranks_unique",
        carrier.loc[carrier["carrier_rank"].notna(), "carrier_rank"].is_unique)
    required_lane_cols = {"shipment_count", "invoiced_freight_usd", "cost_per_kg", "otif_pct",
                          "avg_delay_days", "transit_time_variability_days", "claims_rate_pct",
                          "accessorial_rate_pct", "invoice_exception_pct", "avg_capacity_utilization_pct",
                          "dq_exception_rate_pct", "service_classification", "recommended_action"}
    lane = pd.read_sql("SELECT * FROM lane_scorecard_result", engine)
    add("lane_scorecard_required_metrics", required_lane_cols <= set(lane.columns),
        detail=sorted(required_lane_cols - set(lane.columns)))

    date_dim = pd.read_sql("SELECT date FROM reporting_date_dim ORDER BY date", engine)
    dates = pd.to_datetime(date_dim["date"])
    add("date_dimension_complete", len(dates) == (dates.max() - dates.min()).days + 1, detail=len(dates))
    add("three_root_cause_cases", pd.read_sql("SELECT * FROM root_cause_case_study", engine).shape[0] == 3)
    add("unrated_expected_total_null",
        pd.read_sql("SELECT COUNT(*) n FROM v_freight_audit WHERE rate_status<>'RATED' AND expected_total IS NOT NULL", engine).iloc[0, 0] == 0)
    add("invoice_audit_one_row_per_invoice",
        pd.read_sql("SELECT COUNT(*) n, COUNT(DISTINCT invoice_id) d FROM v_freight_audit", engine).iloc[0].pipe(lambda r: r.n == r.d))
    add("dq_detection_ids_unique",
        pd.read_sql("SELECT COUNT(*) n, COUNT(DISTINCT detected_exception_id) d FROM dq_detected_exception", engine).iloc[0].pipe(lambda r: r.n == r.d))

    expected_exports = ["rpt_dim_date", "rpt_dim_carrier", "rpt_dim_lane", "rpt_dim_product",
                        "rpt_dim_location", "rpt_fact_shipment", "rpt_fact_milestone",
                        "rpt_fact_freight_audit", "rpt_fact_data_quality", "rpt_fact_accrual",
                        "rpt_fact_claim", "rpt_carrier_scorecard", "rpt_lane_scorecard"]
    add("reporting_exports_exist", all((REPORTING / f"{name}.csv").exists() for name in expected_exports))
    add("shipment_export_reconciles", len(pd.read_csv(REPORTING / "rpt_fact_shipment.csv")) == len(ship))

    required_sheets = ["Executive Summary", "Shipment Exceptions", "Carrier Scorecard", "Lane Scorecard",
                       "Freight Audit", "Three-Way Match", "Accrual Report", "Open Claims",
                       "Data Quality", "Metric Definitions"]
    wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
    add("excel_required_sheets", wb.sheetnames == required_sheets, detail=wb.sheetnames)
    ws = wb["Executive Summary"]
    metrics = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(7, ws.max_row + 1)}
    add("excel_shipment_count_reconciles", metrics.get("Shipment count") == len(ship), detail=metrics.get("Shipment count"))
    db_spend = pd.read_sql("SELECT invoiced_freight_usd FROM v_kpi_freight_summary", engine).iloc[0, 0]
    add("excel_freight_spend_reconciles", abs(metrics.get("Freight spend", 0) - db_spend) < 0.01,
        detail=metrics.get("Freight spend"))
    recovery = pd.read_sql("""SELECT ROUND(overcharge_recoverable, 2) AS overcharge_recoverable,
                                      ROUND(duplicate_invoice_exposure, 2) AS duplicate_invoice_exposure,
                                      ROUND(accessorial_recoverable, 2) AS accessorial_recoverable
                                 FROM v_audit_recoverable_summary""", engine).iloc[0]
    expected_overcharge = float(recovery.overcharge_recoverable)
    expected_total_recovery = sum(float(recovery[col]) for col in recovery.index)
    add("excel_recoverable_overcharge_reconciles",
        abs(metrics.get("Recoverable overcharge", 0) - expected_overcharge) < 0.001,
        detail=metrics.get("Recoverable overcharge"))
    add("excel_total_recovery_reconciles",
        abs(metrics.get("Total recoverable exposure", 0) - expected_total_recovery) < 0.001,
        detail=metrics.get("Total recoverable exposure"))
    add("excel_carrier_rows_reconcile", wb["Carrier Scorecard"].max_row - 6 == len(carrier))
    add("excel_lane_rows_reconcile", wb["Lane Scorecard"].max_row - 6 == len(lane))
    wb.close()

    docs = ["phase3_summary.md", "interview_materials.md", "logistics_data_sop.md"]
    add("phase3_documentation_exists", all((PROJECT_ROOT / "documentation" / name).exists() for name in docs))
    result = pd.DataFrame(checks)
    replace_table(engine, result, "rpt_phase3_validation", schema)
    (DATA_PROCESSED / "analytics").mkdir(parents=True, exist_ok=True)
    result.to_csv(DATA_PROCESSED / "analytics" / "rpt_phase3_validation.csv", index=False)
    failures = result[(result["severity"] == "CRITICAL") & (result["status"] == "FAIL")]
    log.info("Phase 3 validation: %d checks, %d failures", len(result), len(failures))
    if len(failures):
        for row in failures.itertuples(index=False):
            log.error("FAILED: %s | %s", row.validation_name, row.detail)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
