"""Validate the existing Phase 3 Excel KPI pack without changing it."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

from common import PROJECT_ROOT, create_project_engine, get_logger, load_config

log = get_logger("validate_excel_kpi_pack")
WORKBOOK = PROJECT_ROOT / "excel" / "logistics_kpi_pack.xlsx"
REPORT = PROJECT_ROOT / "excel" / "kpi_pack_validation.md"
EXPECTED_SHEETS = [
    "Executive Summary",
    "Shipment Exceptions",
    "Carrier Scorecard",
    "Lane Scorecard",
    "Freight Audit",
    "Three-Way Match",
    "Accrual Report",
    "Open Claims",
    "Data Quality",
    "Metric Definitions",
]
SOURCES = {
    "Executive Summary": "Live Phase 3 KPI and control views",
    "Shipment Exceptions": "rpt_fact_shipment exception filter",
    "Carrier Scorecard": "rpt_carrier_scorecard",
    "Lane Scorecard": "rpt_lane_scorecard",
    "Freight Audit": "rpt_fact_freight_audit exception filter",
    "Three-Way Match": "v_three_way_match",
    "Accrual Report": "rpt_fact_accrual",
    "Open Claims": "rpt_fact_claim open filter",
    "Data Quality": "rpt_fact_data_quality",
    "Metric Definitions": "Documented metric-definition registry",
}


def validate() -> tuple[list[dict], dict]:
    cfg = load_config()
    engine = create_project_engine(cfg)
    wb = load_workbook(WORKBOOK, read_only=False, data_only=False)
    checks: list[dict] = []

    def add(name, passed, detail=""):
        checks.append({"check": name, "status": "PASS" if bool(passed) else "FAIL", "detail": str(detail)})

    add("Expected worksheets", wb.sheetnames == EXPECTED_SHEETS, ", ".join(wb.sheetnames))
    formulas = []
    error_tokens = []
    placeholders = []
    for ws in wb.worksheets:
        add(f"{ws.title}: filter", bool(ws.auto_filter.ref), ws.auto_filter.ref)
        add(f"{ws.title}: freeze panes", bool(ws.freeze_panes), ws.freeze_panes)
        add(f"{ws.title}: populated", ws.max_row >= 7 and ws.max_column >= 4, f"{ws.max_row} rows x {ws.max_column} columns")
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    if value.startswith("="):
                        formulas.append(f"{ws.title}!{cell.coordinate}")
                    if any(token in value for token in ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"]):
                        error_tokens.append(f"{ws.title}!{cell.coordinate}")
                    if value.strip().lower() in {"placeholder", "todo", "tbd"}:
                        placeholders.append(f"{ws.title}!{cell.coordinate}")
    add("No formula-error tokens", not error_tokens, error_tokens)
    add("No placeholder rows", not placeholders, placeholders)
    add("Workbook intentionally contains no cell formulas", not formulas, formulas[:10])

    summary = wb["Executive Summary"]
    metrics = {summary.cell(row, 1).value: summary.cell(row, 2).value for row in range(7, summary.max_row + 1)}
    disclosure = str(summary["A3"].value or "").lower()
    add("Executive disclosure visible", "public" in disclosure and "simulated" in disclosure)
    add("Executive Summary is recruiter-readable", len(metrics) >= 20 and "Shipment count" in metrics and "OTIF" in metrics)
    with engine.connect() as connection:
        db_shipment = connection.exec_driver_sql("SELECT shipment_count FROM v_kpi_otif_summary").scalar_one()
        db_spend = connection.exec_driver_sql("SELECT invoiced_freight_usd FROM v_kpi_freight_summary").scalar_one()
        db_overcharge = connection.exec_driver_sql("SELECT ROUND(overcharge_recoverable,2) FROM v_audit_recoverable_summary").scalar_one()
    add("Shipment count reconciles", metrics.get("Shipment count") == db_shipment, metrics.get("Shipment count"))
    add("Freight spend reconciles", abs(metrics.get("Freight spend", 0) - db_spend) < 0.01, metrics.get("Freight spend"))
    add("Recoverable overcharge reconciles", abs(metrics.get("Recoverable overcharge", 0) - db_overcharge) < 0.001, metrics.get("Recoverable overcharge"))
    add("Currency formatting", summary["B15"].number_format == '"$"#,##0.00', summary["B15"].number_format)
    add("Percentage formatting", summary["B10"].number_format == '0.00"%"', summary["B10"].number_format)
    add("Carrier detail row count", wb["Carrier Scorecard"].max_row - 6 == 10, wb["Carrier Scorecard"].max_row - 6)
    add("Lane detail row count", wb["Lane Scorecard"].max_row - 6 == 778, wb["Lane Scorecard"].max_row - 6)
    add("Data-quality detail row count", wb["Data Quality"].max_row - 6 == 4015, wb["Data Quality"].max_row - 6)
    sheet_details = {ws.title: (ws.max_row, ws.max_column, ws.auto_filter.ref, str(ws.freeze_panes)) for ws in wb.worksheets}
    wb.close()
    engine.dispose()
    return checks, sheet_details


def write_report(checks: list[dict], sheet_details: dict) -> None:
    passed = sum(check["status"] == "PASS" for check in checks)
    lines = [
        "# Excel KPI Pack Validation",
        "",
        "Validated workbook: `excel/logistics_kpi_pack.xlsx`",
        "",
        f"**Result:** {passed}/{len(checks)} checks passed. The workbook was reviewed without changing analytical values.",
        "",
        "## Workbook sheets and sources",
        "",
        "| Worksheet | Rows | Columns | Source | Filter / freeze panes |",
        "|---|---:|---:|---|---|",
    ]
    for sheet in EXPECTED_SHEETS:
        rows, columns, filter_ref, freeze = sheet_details[sheet]
        lines.append(f"| {sheet} | {rows:,} | {columns} | {SOURCES[sheet]} | `{filter_ref}` / `{freeze}` |")
    lines.extend([
        "",
        "## Key reconciliation checks",
        "",
        "| Check | Status | Detail |",
        "|---|---|---|",
    ])
    for check in checks:
        lines.append(f'| {check["check"]} | {check["status"]} | {check["detail"]} |')
    lines.extend([
        "",
        "## Refresh procedure",
        "",
        "1. Run `python3 src/run_phase3.py` to refresh the database, reporting CSVs, and workbook.",
        "2. Run `python3 src/run_phase4.py` to regenerate presentation assets and repeat workbook reconciliation.",
        "3. Confirm the Phase 3 and Phase 4 validation reports contain no failed critical checks.",
        "",
        "## Known limitations",
        "",
        "- Shipment patterns are public-source derived; enterprise and finance records are deterministic simulations.",
        "- Expected freight and accrual values are simulated planning/control baselines, not approved company budgets.",
        "- The workbook contains values and conditional formatting rather than cell formulas; refresh is performed by the Python/SQL pipeline.",
        "- The workbook is an Excel management pack, not a substitute for the manually built Power BI report described under `dashboard/`.",
        "",
    ])
    REPORT.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    checks, sheet_details = validate()
    write_report(checks, sheet_details)
    failures = [check for check in checks if check["status"] == "FAIL"]
    if failures:
        for check in failures:
            log.error("Excel validation failed: %s | %s", check["check"], check["detail"])
        return 1
    log.info("Excel KPI pack validation: %d/%d checks pass", len(checks), len(checks))
    return 0


if __name__ == "__main__":
    sys.exit(main())
