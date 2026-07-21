"""Phase 4 portfolio-presentation artifact and reconciliation tests."""

from __future__ import annotations

import csv
import hashlib
import struct
import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

import phase4_visuals  # noqa: E402
import run_phase4  # noqa: E402
from validate_phase4 import CHARTS, DASHBOARDS, DIAGRAMS, DOCS, local_targets  # noqa: E402


def read_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def png_size(path: Path) -> tuple[int, int]:
    data = path.read_bytes()
    assert data[:8] == b"\x89PNG\r\n\x1a\n"
    return struct.unpack(">II", data[16:24])


def visual_paths() -> list[Path]:
    return [PROJECT_ROOT / "dashboard" / "screenshots" / name for name in DASHBOARDS] + [
        PROJECT_ROOT / "documentation" / "charts" / name for name in CHARTS
    ] + [PROJECT_ROOT / "documentation" / "diagrams" / name for name in DIAGRAMS]


def visual_hashes() -> dict[str, str]:
    return {str(path.relative_to(PROJECT_ROOT)): hashlib.sha256(path.read_bytes()).hexdigest() for path in visual_paths()}


def test_six_dashboard_pngs_are_nonempty_and_desktop_resolution():
    paths = [PROJECT_ROOT / "dashboard" / "screenshots" / name for name in DASHBOARDS]
    assert len(paths) == 6
    for path in paths:
        assert path.stat().st_size > 10_000
        assert png_size(path) == (1600, 1000)


def test_eight_standalone_charts_are_nonempty_and_high_resolution():
    paths = [PROJECT_ROOT / "documentation" / "charts" / name for name in CHARTS]
    assert len(paths) == 8
    for path in paths:
        assert path.stat().st_size > 10_000
        assert png_size(path) == (1800, 1100)


def test_three_diagrams_have_png_and_mermaid_sources():
    for name in DIAGRAMS:
        png = PROJECT_ROOT / "documentation" / "diagrams" / name
        mermaid = png.with_suffix(".mmd")
        assert png.stat().st_size > 10_000
        assert png_size(png) == (1800, 900)
        assert "flowchart" in mermaid.read_text(encoding="utf-8")


def test_dashboard_reconciliation_has_all_passes():
    rows = read_csv(PROJECT_ROOT / "dashboard" / "dashboard_metric_reconciliation.csv")
    assert len(rows) == 31
    assert {row["artifact"] for row in rows} == set(DASHBOARDS)
    assert all(row["displayed_value"] == row["source_value"] and row["match"] == "PASS" for row in rows)


def test_chart_source_reconciliation_has_all_passes():
    rows = read_csv(PROJECT_ROOT / "documentation" / "charts" / "chart_source_reconciliation.csv")
    assert len(rows) == 8
    assert {row["artifact"] for row in rows} == set(CHARTS)
    assert all(row["displayed_value"] == row["source_value"] and row["match"] == "PASS" for row in rows)


def test_recruiter_readme_has_required_opening_and_headlines():
    readme = (PROJECT_ROOT / "README.md").read_text(encoding="utf-8")
    assert readme.startswith("# Solar Logistics Control Tower & Freight Audit System")
    for text in ["10,324", "86.88%", "$32.49M", "99.37%", "100%", "$2.48M"]:
        assert text in readme
    assert "modeled within simulated enterprise records" in readme.lower()


def test_readme_image_paths_exist():
    readme = (PROJECT_ROOT / "README.md").read_text(encoding="utf-8")
    images = [target for target in local_targets(readme) if target.endswith(".png")]
    assert len(images) >= 9
    assert all((PROJECT_ROOT / target).exists() for target in images)


def test_no_readme_local_link_is_broken():
    readme = (PROJECT_ROOT / "README.md").read_text(encoding="utf-8")
    missing = [target for target in local_targets(readme) if not (PROJECT_ROOT / target).exists()]
    assert missing == []


def test_required_documentation_exists_and_is_substantive():
    for name in DOCS:
        path = PROJECT_ROOT / "documentation" / name
        assert path.exists()
        assert path.stat().st_size > 500


def test_required_documents_disclose_public_and_simulated_data():
    paths = [
        PROJECT_ROOT / "README.md",
        PROJECT_ROOT / "documentation" / "project_case_study.md",
        PROJECT_ROOT / "documentation" / "recruiter_walkthrough.md",
        PROJECT_ROOT / "documentation" / "interview_demo_script.md",
    ]
    for path in paths:
        text = path.read_text(encoding="utf-8").lower()
        assert "public" in text and "simulat" in text


def test_interview_demo_has_all_timed_sections_and_answer_fields():
    text = (PROJECT_ROOT / "documentation" / "interview_demo_script.md").read_text(encoding="utf-8")
    for minute in range(10):
        assert f"Minute {minute}–{minute + 1}" in text
    assert text.count("**What to show:**") == 10
    assert text.count("**What to say:**") == 10
    assert text.count("**Likely question:**") == 10
    assert text.count("**Strong answer:**") == 10


def test_interview_materials_cover_required_answers_and_star_stories():
    text = (PROJECT_ROOT / "documentation" / "interview_materials.md").read_text(encoding="utf-8")
    for heading in ["30-second answer", "90-second answer", "Two-minute answer", "Four resume bullets", "Five technical", "Five logistics", "Five behavioral"]:
        assert heading in text
    for question in ["Why did you simulate", "Why is precision only 54.94%", "What does 99.37% recall", "How do you know the freight-audit", "What would change with real SAP", "hardest technical issue", "business action would you take first"]:
        assert question.lower() in text.lower()
    assert text.count("**Situation:**") == 5


def test_presentation_outline_has_seven_slides_and_required_fields():
    text = (PROJECT_ROOT / "documentation" / "presentation_outline.md").read_text(encoding="utf-8")
    assert text.count("## Slide ") == 7
    for label in ["**Objective:**", "**Headline:**", "**Visual:**", "**Main messages:**", "**Speaker notes:**", "**Transition:**"]:
        assert text.count(label) == 7


def test_excel_workbook_contains_expected_worksheets_and_controls():
    workbook = PROJECT_ROOT / "excel" / "logistics_kpi_pack.xlsx"
    wb = load_workbook(workbook, read_only=False, data_only=False)
    assert wb.sheetnames == ["Executive Summary", "Shipment Exceptions", "Carrier Scorecard", "Lane Scorecard", "Freight Audit", "Three-Way Match", "Accrual Report", "Open Claims", "Data Quality", "Metric Definitions"]
    for ws in wb.worksheets:
        assert ws.auto_filter.ref
        assert ws.freeze_panes
    wb.close()


def test_excel_headline_values_reconcile():
    wb = load_workbook(PROJECT_ROOT / "excel" / "logistics_kpi_pack.xlsx", read_only=True, data_only=False)
    ws = wb["Executive Summary"]
    metrics = {ws.cell(row, 1).value: ws.cell(row, 2).value for row in range(7, ws.max_row + 1)}
    assert metrics["Shipment count"] == 10324
    assert metrics["OTIF"] == 86.88
    assert metrics["Recoverable overcharge"] == 428053.23
    assert metrics["Total recoverable exposure"] == 2477840.98
    assert metrics["Critical recall"] == 100
    wb.close()


def test_excel_validation_report_records_full_pass():
    text = (PROJECT_ROOT / "excel" / "kpi_pack_validation.md").read_text(encoding="utf-8")
    assert "44/44 checks passed" in text
    assert "## Refresh procedure" in text and "## Known limitations" in text


def test_phase4_generation_is_idempotent():
    before = visual_hashes()
    phase4_visuals.generate_all()
    after = visual_hashes()
    assert before == after


def test_phase4_hash_manifest_records_all_passes():
    rows = read_csv(PROJECT_ROOT / "documentation" / "phase4_artifact_hashes.csv")
    assert len(rows) >= 20
    assert all(row["first_sha256"] == row["second_sha256"] and row["match"] == "PASS" for row in rows)


def test_repository_baseline_guard_accepts_current_history():
    assert run_phase4.verify_repository() == 0
