"""Phase 3 analytics, audit, scorecard, reporting, and artifact regression tests."""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, inspect

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from build_scorecards import build_carrier_scorecard, normalize  # noqa: E402
from common import database_url, load_config  # noqa: E402
from gen_common import adapt_sql_for_sqlite, split_sql_statements  # noqa: E402


@pytest.fixture(scope="module")
def engine():
    return create_engine(database_url(load_config()))


def q(engine, sql):
    return pd.read_sql(sql, engine)


def test_required_phase3_views_exist(engine):
    views = set(inspect(engine).get_view_names())
    assert {"v_dq_detected", "v_kpi_otif_summary", "v_freight_audit",
            "v_three_way_match", "v_accrual_summary", "v_carrier_metrics",
            "v_lane_metrics", "rpt_fact_shipment", "rpt_carrier_scorecard",
            "rpt_lane_scorecard"} <= views


def test_critical_manifest_detection_is_complete(engine):
    p = q(engine, "SELECT * FROM rpt_dq_detection_performance WHERE manifest_count>0 AND severity='CRITICAL'")
    assert p["false_negative_count"].sum() == 0
    assert p["true_positive_count"].sum() == 270


def test_overall_detection_recall_target(engine):
    p = q(engine, "SELECT * FROM rpt_dq_detection_performance WHERE manifest_count>0")
    recall = p["true_positive_count"].sum() / (p["true_positive_count"].sum() + p["false_negative_count"].sum())
    assert recall >= 0.95
    assert round(recall * 100, 2) == 99.37


def test_detection_precision_reconciles(engine):
    p = q(engine, "SELECT * FROM rpt_dq_detection_performance WHERE manifest_count>0")
    precision = p["true_positive_count"].sum() / (p["true_positive_count"].sum() + p["false_positive_count"].sum())
    assert round(precision * 100, 2) == 54.94


def test_expired_rate_gap_is_documented(engine):
    row = q(engine, "SELECT * FROM rpt_dq_detection_performance WHERE exception_type='expired_rate_card'").iloc[0]
    assert row.recall_pct == 92.0
    assert "12" in row.notes and "share" in row.notes.lower()


def test_otif_reconciles_to_analytics(engine):
    summary = q(engine, "SELECT * FROM v_kpi_otif_summary").iloc[0]
    ship = q(engine, "SELECT * FROM analytics_shipment")
    delivered = ship[ship["is_delivered"] == 1]
    assert summary.shipment_count == len(ship) == 10324
    assert summary.delivered_count == len(delivered) == 10012
    assert summary.otif_pct == round(delivered["otif_flag"].mean() * 100, 2)


def test_git_count_and_age(engine):
    git = q(engine, "SELECT * FROM analytics_shipment WHERE git_flag=1")
    assert len(git) == 312
    assert git["git_age_days"].min() >= 0
    assert (git["overdue_git_flag"].isin([0, 1])).all()


def test_transit_time_reconciles(engine):
    got = q(engine, "SELECT AVG(actual_transit_days) value FROM analytics_shipment WHERE is_delivered=1").iloc[0, 0]
    source = q(engine, "SELECT actual_ship_date, actual_delivery_date FROM fact_shipment WHERE shipment_status='DELIVERED'")
    expected = (pd.to_datetime(source.actual_delivery_date) - pd.to_datetime(source.actual_ship_date)).dt.days.mean()
    assert abs(got - expected) < 1e-9


def test_freight_cost_per_kg_reconciles(engine):
    row = q(engine, "SELECT * FROM v_kpi_freight_by_carrier WHERE carrier_id='CAR10'").iloc[0]
    raw = q(engine, "SELECT SUM(i.invoice_total) spend, SUM(s.shipment_weight_kg) weight FROM fact_freight_invoice i JOIN fact_shipment s ON i.shipment_id=s.shipment_id WHERE s.carrier_id='CAR10'").iloc[0]
    assert abs(row.cost_per_kg - round(raw.spend / raw.weight, 4)) < 1e-9


def test_pod_compliance_reconciles(engine):
    pod = q(engine, "SELECT * FROM v_kpi_pod_compliance").iloc[0]
    assert pod.delivered_count == 10012
    assert pod.pod_count + pod.missing_pod_count == pod.delivered_count
    assert pod.pod_compliance_pct == 96.0


def test_claims_rate_reconciles(engine):
    claims = q(engine, "SELECT * FROM v_kpi_claims").iloc[0]
    assert claims.claim_count == 411
    assert claims.claims_rate_pct == round(411 / 10012 * 100, 3)


def test_expected_freight_one_row_per_shipment(engine):
    row = q(engine, "SELECT COUNT(*) n, COUNT(DISTINCT shipment_id) d FROM v_expected_freight").iloc[0]
    assert row.n == row.d == 10324


def test_rate_selection_respects_effective_dates(engine):
    bad = q(engine, "SELECT COUNT(*) n FROM v_expected_freight e JOIN dim_rate_card r ON e.rate_id=r.rate_id WHERE e.actual_ship_date NOT BETWEEN r.effective_start_date AND r.effective_end_date").iloc[0, 0]
    assert bad == 0


def test_minimum_charge_is_enforced(engine):
    bad = q(engine, "SELECT COUNT(*) n FROM v_expected_freight WHERE rate_status='RATED' AND expected_base < minimum_charge").iloc[0, 0]
    assert bad == 0


def test_expected_fuel_calculation(engine):
    rated = q(engine, "SELECT expected_base, expected_fuel, fuel_percentage FROM v_expected_freight WHERE rate_status='RATED'")
    assert np.allclose(rated.expected_fuel, (rated.expected_base * rated.fuel_percentage).round(2), atol=0.01)


def test_accessorial_expected_amount_reconciles(engine):
    sample = q(engine, "SELECT invoice_id, expected_accessorial FROM v_freight_audit WHERE expected_accessorial>0 LIMIT 100")
    bands = q(engine, "SELECT * FROM meta_accessorial_band")
    acc = q(engine, "SELECT * FROM fact_accessorial_charge")
    valid = acc.merge(bands, on="charge_type")
    valid = valid[(valid.contractually_allowed_flag == 1) & (valid.supporting_document_flag == 1) & (valid.charge_amount <= valid.max_allowed)]
    expected = valid.groupby("invoice_id").charge_amount.sum()
    assert all(abs(r.expected_accessorial - expected.get(r.invoice_id, 0)) < 0.01 for r in sample.itertuples())


def test_expected_invoice_total_components(engine):
    rated = q(engine, "SELECT * FROM v_freight_audit WHERE expected_total IS NOT NULL")
    calc = rated.expected_base + rated.expected_fuel.fillna(0) + rated.expected_accessorial.fillna(0) + rated.tax_amount
    assert np.allclose(rated.expected_total, calc, atol=0.01)


def test_unrated_invoices_do_not_become_zero_expectations(engine):
    bad = q(engine, "SELECT COUNT(*) n FROM v_freight_audit WHERE rate_status<>'RATED' AND expected_total IS NOT NULL").iloc[0, 0]
    assert bad == 0


def test_overcharge_and_undercharge_amounts_nonnegative(engine):
    audit = q(engine, "SELECT overcharge_amount, undercharge_amount FROM v_freight_audit")
    assert (audit.overcharge_amount >= 0).all()
    assert (audit.undercharge_amount >= 0).all()


def test_duplicate_invoice_detection(engine):
    dup = q(engine, "SELECT COUNT(*) n FROM v_freight_audit WHERE audit_status='DUPLICATE_INVOICE'").iloc[0, 0]
    assert dup == 120


def test_currency_mismatch_detection(engine):
    wrong = q(engine, "SELECT COUNT(*) n FROM v_freight_audit WHERE audit_status='INCORRECT_CURRENCY'").iloc[0, 0]
    assert wrong == 43


def test_null_shipment_carrier_not_called_carrier_mismatch(engine):
    bad = q(engine, "SELECT COUNT(*) n FROM v_freight_audit WHERE shipment_carrier IS NULL AND audit_status='CARRIER_MISMATCH'").iloc[0, 0]
    assert bad == 0


def test_three_way_match_counts(engine):
    counts = q(engine, "SELECT overall_match_status, COUNT(*) n FROM v_three_way_match GROUP BY overall_match_status").set_index("overall_match_status").n.to_dict()
    assert counts == {"BLOCK_PAYMENT": 219, "MATCHED": 7275, "MATCHED_WITH_WARNING": 155,
                      "MISSING_RECORD": 60, "REVIEW_REQUIRED": 2233}


def test_three_way_quantity_compares_shipment_quantity(engine):
    exc = q(engine, "SELECT COUNT(*) n FROM v_three_way_match WHERE quantity_match_status='EXCEPTION'").iloc[0, 0]
    direct = q(engine, "SELECT COUNT(*) n FROM fact_freight_invoice i JOIN fact_shipment s ON i.shipment_id=s.shipment_id WHERE s.delivered_quantity>s.planned_quantity").iloc[0, 0]
    assert exc == direct == 82


def test_released_accrual_eligibility(engine):
    bad = q(engine, "SELECT COUNT(*) n FROM fact_accrual WHERE accrual_status='RELEASED' AND (invoice_received_flag<>1 OR actual_invoice_cost IS NULL)").iloc[0, 0]
    assert bad == 0


def test_accrual_summary_reconciles(engine):
    summary = q(engine, "SELECT * FROM v_accrual_summary").iloc[0]
    raw = q(engine, "SELECT * FROM fact_accrual")
    assert summary.accrual_count == len(raw) == 10324
    assert abs(summary.open_accrual_balance - raw.loc[raw.accrual_status != "RELEASED", "expected_freight_cost"].sum()) < 0.01


def test_carrier_score_weights_sum_to_one(engine):
    assert abs(q(engine, "SELECT SUM(weight) value FROM meta_scorecard_weight").iloc[0, 0] - 1.0) < 1e-9


def test_normalization_directionality():
    values = pd.Series([1.0, 2.0, 3.0])
    assert normalize(values, True).tolist() == [0.0, 50.0, 100.0]
    assert normalize(values, False).tolist() == [100.0, 50.0, 0.0]


def test_carrier_weighting_and_low_volume_ranking():
    cfg = load_config()
    rows = []
    for cid, n, boost in [("A", 100, 10), ("B", 100, 0), ("C", 10, 100)]:
        rows.append({"carrier_id": cid, "carrier_name": cid, "shipment_count": n, "delivered_count": n,
                     "otif_pct": 80 + boost, "transit_reliability_pct": 70 + boost,
                     "pod_compliance_pct": 90 + min(boost, 10), "avg_delay_days": 0,
                     "avg_actual_transit_days": 5, "invoiced_freight_usd": 1000,
                     "cost_per_kg": 5 - min(boost, 4) / 10, "invoice_count": n,
                     "invoice_accuracy_pct": 85 + min(boost, 10), "claim_count": 1,
                     "claims_rate_pct": 2 - min(boost, 1)})
    out = build_carrier_scorecard(pd.DataFrame(rows), cfg).set_index("carrier_id")
    assert out.loc["A", "total_score"] > out.loc["B", "total_score"]
    assert pd.isna(out.loc["C", "carrier_rank"])
    assert out.loc["C", "classification"] == "Insufficient volume"


def test_carrier_rank_is_unique_for_eligible(engine):
    sc = q(engine, "SELECT * FROM carrier_scorecard_result WHERE carrier_rank IS NOT NULL")
    assert sc.carrier_rank.is_unique
    assert sc.total_score.between(0, 100).all()


def test_lane_scorecard_has_required_metrics(engine):
    lane = q(engine, "SELECT * FROM lane_scorecard_result")
    required = {"shipment_count", "invoiced_freight_usd", "cost_per_kg", "otif_pct",
                "avg_delay_days", "transit_time_variability_days", "claims_rate_pct",
                "accessorial_rate_pct", "invoice_exception_pct", "avg_capacity_utilization_pct",
                "dq_exception_rate_pct", "service_classification", "recommended_action"}
    assert required <= set(lane.columns)


def test_lane_insufficient_volume_gate(engine):
    lane = q(engine, "SELECT * FROM lane_scorecard_result")
    low = lane[lane.shipment_count < load_config()["phase3"]["lane_scorecard"]["min_shipments"]]
    assert (low.service_classification == "Insufficient volume").all()


def test_power_bi_shipment_view_reconciles(engine):
    assert q(engine, "SELECT COUNT(*) n FROM rpt_fact_shipment").iloc[0, 0] == q(engine, "SELECT COUNT(*) n FROM fact_shipment").iloc[0, 0]


def test_power_bi_audit_view_reconciles(engine):
    assert q(engine, "SELECT COUNT(*) n FROM rpt_fact_freight_audit").iloc[0, 0] == q(engine, "SELECT COUNT(*) n FROM fact_freight_invoice").iloc[0, 0]


def test_date_dimension_is_complete(engine):
    dates = pd.to_datetime(q(engine, "SELECT date FROM rpt_dim_date ORDER BY date").date)
    assert len(dates) == (dates.max() - dates.min()).days + 1
    assert dates.diff().dropna().eq(pd.Timedelta(days=1)).all()


def test_reporting_exports_exist_and_reconcile(engine):
    reporting = PROJECT_ROOT / "data" / "processed" / "reporting"
    assert len(pd.read_csv(reporting / "rpt_fact_shipment.csv")) == 10324
    assert len(pd.read_csv(reporting / "rpt_fact_freight_audit.csv")) == 9942
    assert len(pd.read_csv(reporting / "rpt_carrier_scorecard.csv")) == len(q(engine, "SELECT * FROM carrier_scorecard_result"))


def test_excel_required_sheets():
    wb = load_workbook(PROJECT_ROOT / "excel" / "logistics_kpi_pack.xlsx", read_only=True)
    assert wb.sheetnames == ["Executive Summary", "Shipment Exceptions", "Carrier Scorecard",
                             "Lane Scorecard", "Freight Audit", "Three-Way Match", "Accrual Report",
                             "Open Claims", "Data Quality", "Metric Definitions"]
    wb.close()


def test_excel_summary_reconciles_sql(engine):
    wb = load_workbook(PROJECT_ROOT / "excel" / "logistics_kpi_pack.xlsx", read_only=True, data_only=False)
    ws = wb["Executive Summary"]
    metrics = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(7, ws.max_row + 1)}
    assert metrics["Shipment count"] == 10324
    assert abs(metrics["Freight spend"] - q(engine, "SELECT invoiced_freight_usd FROM v_kpi_freight_summary").iloc[0, 0]) < 0.01
    recovery = q(engine, """SELECT ROUND(overcharge_recoverable, 2) AS overcharge_recoverable,
                                    ROUND(duplicate_invoice_exposure, 2) AS duplicate_invoice_exposure,
                                    ROUND(accessorial_recoverable, 2) AS accessorial_recoverable
                               FROM v_audit_recoverable_summary""").iloc[0]
    assert metrics["Recoverable overcharge"] == float(recovery.overcharge_recoverable)
    assert metrics["Total recoverable exposure"] == sum(float(recovery[col]) for col in recovery.index)
    assert metrics["Critical recall"] == 100
    wb.close()


def test_excel_detail_row_counts_reconcile(engine):
    wb = load_workbook(PROJECT_ROOT / "excel" / "logistics_kpi_pack.xlsx", read_only=True)
    assert wb["Carrier Scorecard"].max_row - 6 == q(engine, "SELECT COUNT(*) n FROM carrier_scorecard_result").iloc[0, 0]
    assert wb["Lane Scorecard"].max_row - 6 == q(engine, "SELECT COUNT(*) n FROM lane_scorecard_result").iloc[0, 0]
    assert wb["Data Quality"].max_row - 6 == q(engine, "SELECT COUNT(*) n FROM dq_detected_exception").iloc[0, 0]
    wb.close()


def test_full_shipment_lineage(engine):
    row = q(engine, "SELECT COUNT(*) n, COUNT(DISTINCT source_record_id) d FROM rpt_fact_shipment").iloc[0]
    stg = q(engine, "SELECT COUNT(*) n FROM stg_shipment").iloc[0, 0]
    assert row.n == row.d == stg == 10324


def test_root_cause_outputs_are_complete(engine):
    cases = q(engine, "SELECT * FROM root_cause_case_study")
    assert len(cases) == 3
    for col in ["problem", "evidence", "drill_down", "possible_cause", "business_impact",
                "corrective_action", "follow_up_kpi", "business_owner"]:
        assert cases[col].notna().all() and cases[col].str.len().gt(0).all()


def test_sql_splitter_preserves_semicolon_inside_string():
    sql = "CREATE TABLE x(a TEXT); INSERT INTO x VALUES ('a;b'); -- ignored; comment\nSELECT 1;"
    statements = split_sql_statements(sql)
    assert len(statements) == 3
    assert "'a;b'" in statements[1]


def test_sql_splitter_handles_escaped_quote():
    statements = split_sql_statements("SELECT 'it''s;valid'; SELECT 2;")
    assert statements == ["SELECT 'it''s;valid'", "SELECT 2"]


def test_sqlite_adaptation_removes_schema_qualifier():
    sql = adapt_sql_for_sqlite("CREATE VIEW sunlog.v AS SELECT * FROM sunlog.t;", "sunlog")
    assert "sunlog." not in sql


def test_phase1_ddl_is_compatible_with_phase2_shipment_schema(engine):
    from load_database import run_ddl
    run_ddl(engine, load_config()["database"]["schema"])
    indexes = {idx["name"] for idx in inspect(engine).get_indexes("fact_shipment")}
    assert "ix_shipment_dates" in indexes


def test_dax_contains_required_safe_division_measures():
    dax = (PROJECT_ROOT / "dashboard" / "measures.dax").read_text(encoding="utf-8")
    for name in ["OTIF Percentage", "GIT Count", "Freight Variance", "Potential Overcharge",
                 "Invoice Accuracy", "Claims Rate", "Open Accrual", "Data Quality Score",
                 "Detection Precision", "Detection Recall", "Carrier Score"]:
        assert f"{name} :=" in dax
    assert dax.count("DIVIDE") >= 8


def test_validation_report_has_no_failures(engine):
    val = q(engine, "SELECT * FROM rpt_phase3_validation")
    assert len(val) >= 20
    assert (val.status == "PASS").all()


def test_core_phase3_load_is_idempotent(engine):
    import load_phase3
    before = q(engine, "SELECT COUNT(*) n FROM dq_detected_exception").iloc[0, 0]
    assert load_phase3.main() == 0
    assert load_phase3.main() == 0
    after = q(engine, "SELECT COUNT(*) n FROM dq_detected_exception").iloc[0, 0]
    assert before == after == 4015
